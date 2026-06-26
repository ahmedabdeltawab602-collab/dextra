"""Smart, transparent data loader for dextra - Phase 11 (the entry layer).

See ``LOADER_PHILOSOPHY.md`` and ``LOADER_SPEC_11_1.md``. Stage 11.1 covers
delimited text (csv/tsv) plus a typed pass-through for in-memory frames;
stage 11.2 adds Excel (xlsx/xlsm via openpyxl, lazily): sheet listing and
selection, data-block detection, multi-row headers, values-not-formulas;
stage 11.3a adds parquet (typed pass-through via a lazy engine) and
JSON/NDJSON (records array / one-object-per-line, nested values disclosed);
stage 11.3b adds safe SQL: parametrised statements only (one statement,
values bound via ``sql_params=``), SQLite files opened read-only, and a
default row guard against runaway result sets.

Governing principle: *transparency scales with uncertainty*. Confident parses
load in one line and are fully disclosed; ambiguous decisions are flagged and,
under the chosen policy, warned about / raised / returned-as-a-plan rather than
guessed silently. Every load emits a JSON-safe, replayable **load plan** (the
unified-contract ``params`` artifact). The source is never modified.
"""

from __future__ import annotations

import csv
import datetime as _dt
import hashlib
import io
import json
import os
import re
import warnings
from typing import Optional

import numpy as np
import pandas as pd

from ._utils import _ensure_pandas, append_audit, get_variable_name, json_safe, now_iso
from ._version import __version__

# ---------------------------------------------------------------------------
# Exceptions & warning
# ---------------------------------------------------------------------------


class DextraLoaderError(Exception):
    """Base class for loader errors."""


class LoaderSecurityError(DextraLoaderError):
    """Raised when a source is refused for security reasons (e.g. pickle)."""


class LoaderAmbiguityError(DextraLoaderError):
    """Raised under ``on_ambiguous='raise'`` when a parse decision is ambiguous."""


class LoaderReplayError(DextraLoaderError):
    """Raised when a replay plan no longer matches its source."""


class LoaderAbort(DextraLoaderError):
    """Raised when an interactive user aborts the load."""


class DextraLoaderWarning(UserWarning):
    """Warning emitted under ``on_ambiguous='warn'`` listing ambiguous decisions."""


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_PICKLE_EXT = {".pkl", ".pickle"}
_TSV_EXT = {".tsv", ".tab"}
_EXCEL_EXT = {".xlsx", ".xlsm"}
_XLS_EXT = {".xls"}
_PARQUET_EXT = {".parquet", ".pq"}
_JSON_EXT = {".json", ".jsonl", ".ndjson"}
_SQLITE_EXT = {".db", ".sqlite", ".sqlite3"}
_SQL_ROW_GUARD = 1_000_000  # default row guard for SQL results
_BOOL_TOKENS = {"true", "false", "yes", "no", "y", "n", "t", "f", "1", "0"}
_TRUE_TOKENS = {"true", "yes", "y", "t", "1"}
_HIGH_RISK_RE = re.compile(r"(^id$|_id$|^key$|key$|target|label|^y$)", re.IGNORECASE)
_DATE_FORMATS = ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y",
                 "%Y/%m/%d", "%d-%m-%Y")
_CONFIRMED = "confirmed"
_AMBIGUOUS = "ambiguous"
_HIGH_RISK = "ambiguous-high-risk"
_PARSE_ACCEPT = 0.95  # min parse-rate to accept a non-text dtype


# ---------------------------------------------------------------------------
# Source reading
# ---------------------------------------------------------------------------

def _read_bytes(source):
    """Return ``(raw_bytes, source_meta)`` from a path / file-like object."""
    if isinstance(source, (str, os.PathLike)):
        path = os.fspath(source)
        ext = os.path.splitext(path)[1].lower()
        if ext in _PICKLE_EXT:
            raise LoaderSecurityError(
                f"refusing to load pickle {path!r}: pickle can execute "
                "arbitrary code, so dextra never auto-loads it. Convert it to "
                "CSV/Parquet, or read it yourself with pandas.read_pickle if "
                "you trust its origin.")
        with open(path, "rb") as fh:
            raw = fh.read()
        try:
            stat = os.stat(path)
            size, mtime = int(stat.st_size), float(stat.st_mtime)
        except OSError:
            size, mtime = None, None
        meta = {"name": os.path.basename(path),
                "sha256": hashlib.sha256(raw).hexdigest(),
                "size": size, "mtime": mtime}
        return raw, meta
    if hasattr(source, "read"):
        data = source.read()
        if isinstance(data, str):
            data = data.encode("utf-8")
        meta = {"name": getattr(source, "name", "<stream>"),
                "sha256": hashlib.sha256(data).hexdigest(),
                "size": len(data), "mtime": None}
        return data, meta
    raise TypeError(
        f"load: source must be a path, file-like object, or DataFrame; "
        f"got {type(source).__name__}.")


# ---------------------------------------------------------------------------
# Detection: encoding / decode / delimiter / header
# ---------------------------------------------------------------------------

def _arabic_runs(text: str) -> bool:
    """True when non-ASCII chars are mostly Arabic-block AND form runs.

    Legacy-Arabic bytes decoded with cp1256 yield consecutive Arabic letters
    (whole words), while European accented text (latin-1 / cp125x) yields
    isolated high chars between ASCII letters -- adjacency separates the two
    even though both occupy the same high-byte range.
    """
    idx = [i for i, ch in enumerate(text) if ord(ch) > 127]
    if not idx:
        return False
    arabic = sum(1 for i in idx if 0x0600 <= ord(text[i]) <= 0x06FF)
    adjacent = sum(
        1 for k, i in enumerate(idx)
        if (k > 0 and idx[k - 1] == i - 1)
        or (k + 1 < len(idx) and idx[k + 1] == i + 1))
    return (arabic / len(idx) >= 0.8) and (adjacent / len(idx) >= 0.5)


def _detect_encoding(sample: bytes, forced: Optional[str],
                     truncated: bool = False):
    """Pick an encoding for ``sample``; only certainty earns ``confirmed``.

    Order: user override > BOM > strict utf-8 > guesses. Everything below
    the strict-utf-8 line is a guess and is reported ``ambiguous`` so the
    ``on_ambiguous`` policy surfaces it (eval M-6: a detector's first guess
    used to be stamped ``confirmed``, turning cp1256 -- and even valid
    utf-8 -- Arabic into mojibake silently). ``encoding=`` is the
    deterministic way out.
    """
    if forced:
        return forced, _CONFIRMED, f"user-specified ({forced})"
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig", _CONFIRMED, "BOM: utf-8"
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16", _CONFIRMED, "BOM: utf-16"
    if b"\x00" not in sample:  # NUL bytes suggest BOM-less utf-16 -> guess
        try:
            sample.decode("utf-8")
            return "utf-8", _CONFIRMED, "strict utf-8 decode of sample"
        except UnicodeDecodeError as exc:
            if truncated and exc.start >= len(sample) - 3:
                # a multi-byte char split at the sample cut, not bad bytes;
                # _decode_full still strict-decodes the full payload after
                return ("utf-8", _CONFIRMED,
                        "strict utf-8 decode of sample (boundary-split tail)")
    try:
        if _arabic_runs(sample.decode("cp1256")):
            return ("cp1256", _AMBIGUOUS,
                    "guessed cp1256 (legacy Arabic byte runs); pass "
                    "encoding= to confirm")
    except UnicodeDecodeError:
        pass
    try:
        from charset_normalizer import from_bytes  # lazy `io` extra
        best = from_bytes(sample).best()
        if best is not None and best.encoding:
            enc = ("utf-8" if best.encoding.replace("_", "-") == "utf-8"
                   else best.encoding)
            return (enc, _AMBIGUOUS,
                    f"charset-normalizer guess: {enc}; pass encoding= to "
                    "confirm")
    except Exception:  # noqa: BLE001 - fall back to stdlib probing
        pass
    try:
        sample.decode("cp1256")
        return "cp1256", _AMBIGUOUS, "fallback: cp1256 decoded (utf-8 failed)"
    except UnicodeDecodeError:
        return "latin-1", _AMBIGUOUS, "fallback: latin-1 (lossless byte map)"


def _decode_full(raw: bytes, enc: str, conf: str, reason: str):
    """Decode the whole payload; downgrade to latin-1 if strict decode fails."""
    try:
        return raw.decode(enc), conf, reason
    except (UnicodeDecodeError, LookupError):
        return (raw.decode("latin-1"), _AMBIGUOUS,
                f"{reason}; strict decode failed beyond sample -> latin-1")


def _detect_delimiter(sample_text: str, forced: Optional[str], exclude=()):
    if forced:
        return forced, _CONFIRMED, f"user-specified ({forced!r})"
    candidates = [c for c in (",", ";", "\t", "|") if c not in exclude]
    try:
        from clevercsv import Sniffer as _CleverSniffer  # lazy `io` extra
        dialect = _CleverSniffer().sniff(sample_text)
        if (dialect is not None and dialect.delimiter
                and dialect.delimiter not in exclude):
            return dialect.delimiter, _CONFIRMED, "clevercsv sniffer"
    except Exception:  # noqa: BLE001 - fall back to stdlib csv.Sniffer
        pass
    head = "\n".join(sample_text.splitlines()[:10])
    sep = None
    if candidates:
        try:
            sep = csv.Sniffer().sniff(head, delimiters="".join(candidates)).delimiter
        except csv.Error:
            sep = None
    if sep is None:
        present = [c for c in candidates if head.count(c) > 0]
        if not present:
            for cand in ("\t", "|", ";", ",", "\x01"):
                if cand not in exclude and cand not in head:
                    return cand, _CONFIRMED, "single column (no delimiter detected)"
            return "\x01", _CONFIRMED, "single column"
        sep = max(present, key=head.count)
    # Stability: quote-aware field counts; rows before the first modal-width
    # row are treated as preamble (the header detector will skip them too).
    sample30 = "\n".join(sample_text.splitlines()[:30])
    rows = [r for r in _parse_rows(sample30, sep) if r]
    counts = [len(r) for r in rows]
    if not counts:
        return sep, _AMBIGUOUS, "no parsable rows in the sample"
    modal = max(set(counts), key=counts.count)
    idx = counts.index(modal)
    stable = all(c == modal for c in counts[idx:])
    conf = _CONFIRMED if stable else _AMBIGUOUS
    if stable:
        reason = ("delimiter field count stable" if idx == 0 else
                  f"delimiter field count stable after {idx} preamble row(s)")
    else:
        reason = "delimiter field count varies across rows"
    return sep, conf, reason


def _detect_header(rows, forced: Optional[int]):
    """Pick the header row index among parsed ``rows`` (list of field lists)."""
    if forced is not None:
        return int(forced), _CONFIRMED, f"user-specified (row {forced})"
    probe = [r for r in rows[:20]]
    if not probe:
        return 0, _AMBIGUOUS, "empty input -> row 0"
    counts = [len(r) for r in probe]
    modal = max(set(counts), key=counts.count)
    for i, r in enumerate(probe):
        if len(r) != modal:
            continue
        cells = [c for c in r if str(c).strip() != ""]
        if not cells:
            continue
        non_numeric = sum(0 if _looks_numeric(c) else 1 for c in cells)
        if non_numeric >= max(1, len(cells) // 2):
            stable = counts.count(modal) >= max(2, len(counts) - i - 1)
            conf = _CONFIRMED if (i == 0 or stable) else _AMBIGUOUS
            reason = (f"row {i}: {modal} header-like fields"
                      + ("" if conf == _CONFIRMED else "; preamble field counts vary"))
            return i, conf, reason
    return 0, _AMBIGUOUS, "no clear header row -> row 0"


def _looks_numeric(token: str) -> bool:
    t = str(token).strip().replace(",", "").replace(" ", "")
    if t == "":
        return False
    try:
        float(t)
        return True
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# Per-column type inference (measured)
# ---------------------------------------------------------------------------

def _nonnull(s: pd.Series) -> pd.Series:
    return s[s.notna() & (s.astype(str).str.strip() != "")]

def _try_datetime(s: pd.Series):
    base = _nonnull(s)
    if base.empty:
        return None, 0.0
    text = base.astype(str).str.strip()
    if text.str.fullmatch(r"[0-9]+").all():
        # B-1 guard: pure-digit strings ("001", "20240102") carry no date
        # separator -- parsing them as datetime invents meaning (year 1).
        # Leave them to the numeric / text paths.
        return None, 0.0
    for fmt in _DATE_FORMATS:
        conv = pd.to_datetime(base, format=fmt, errors="coerce")
        if conv.notna().mean() >= _PARSE_ACCEPT:
            return pd.to_datetime(s, format=fmt, errors="coerce"), float(conv.notna().mean())
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        conv = pd.to_datetime(base, errors="coerce")
    rate = float(conv.notna().mean())
    if rate >= _PARSE_ACCEPT:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            full = pd.to_datetime(s, errors="coerce")
        return full, rate
    return None, rate


def _try_numeric(s: pd.Series, decimal: str, thousands: Optional[str]):
    base = _nonnull(s)
    if base.empty:
        return None, 0.0, None
    cleaned = base.astype(str).str.strip()
    suggest = None
    if (cleaned.str.endswith("%")).mean() > 0.5:
        suggest = "% stripped (value kept as written, not divided by 100)"
    cleaned = cleaned.str.replace("%", "", regex=False)
    cleaned = cleaned.str.replace(r"^[\$€£¥]", "", regex=True)
    if thousands:
        cleaned = cleaned.str.replace(thousands, "", regex=False)
    if decimal and decimal != ".":
        cleaned = cleaned.str.replace(decimal, ".", regex=False)
    conv = pd.to_numeric(cleaned, errors="coerce")
    rate = float(conv.notna().mean())
    if rate >= _PARSE_ACCEPT:
        full = s.astype(str).str.strip().str.replace("%", "", regex=False)
        full = full.str.replace(r"^[\$€£¥]", "", regex=True)
        if thousands:
            full = full.str.replace(thousands, "", regex=False)
        if decimal and decimal != ".":
            full = full.str.replace(decimal, ".", regex=False)
        out = pd.to_numeric(full, errors="coerce")
        return out, rate, suggest
    return None, rate, suggest


def _try_bool(s: pd.Series):
    base = _nonnull(s).astype(str).str.strip().str.lower()
    if base.empty:
        return None, 0.0
    if set(base.unique()).issubset(_BOOL_TOKENS):
        mapped = s.astype(str).str.strip().str.lower().map(
            lambda v: True if v in _TRUE_TOKENS else (False if v in _BOOL_TOKENS else np.nan))
        return mapped.astype("boolean"), 1.0
    return None, 0.0


def _infer_column(name: str, s: pd.Series, parse_dates: bool,
                  decimal: str, thousands: Optional[str]):
    """Return ``(typed_series, column_plan_dict)``."""
    base = _nonnull(s)
    n_base = int(base.shape[0])

    typed, dtype, rate, suggest = s, "object", 1.0, None

    if n_base:
        text = base.astype(str).str.strip()
        if (text.str.fullmatch(r"[0-9]+").all()
                and bool(((text.str.len() > 1)
                          & text.str.startswith("0")).any())):
            # B-1 guard: pure-digit values with a leading zero are identifiers
            # (order ids, zipcodes, phones) -- int/datetime coercion would
            # silently drop the zeros. Keep text, never CONFIRMED, so the
            # on_ambiguous policy (warn/raise/plan) surfaces the decision.
            conf = (_HIGH_RISK if _HIGH_RISK_RE.search(str(name))
                    else _AMBIGUOUS)
            col_plan = {"dtype": "object", "coerced_from": "object",
                        "parse_rate": 1.0, "n_failed": 0,
                        "confidence": conf,
                        "reason": "leading-zero identifier kept as text",
                        "suggest": None}
            return s, col_plan

    if parse_dates:
        conv, r = _try_datetime(s)
        if conv is not None:
            typed, dtype, rate = conv, "datetime64[ns]", r
    if dtype == "object":
        conv, r, sug = _try_numeric(s, decimal, thousands)
        if conv is not None:
            typed, dtype, rate, suggest = conv, "float64", r, sug
    if dtype == "object":
        conv, r = _try_bool(s)
        if conv is not None:
            typed, dtype, rate = conv, "boolean", r
    if dtype == "object" and n_base and (base.nunique() / n_base) <= 0.5:
        suggest = "category"

    if dtype != "object":
        dtype = str(typed.dtype)

    n_failed = 0
    if dtype != "object" and n_base:
        n_failed = int(round((1.0 - rate) * n_base))

    all_nan = bool(dtype != "object" and typed.notna().sum() == 0)
    high_risk = bool(_HIGH_RISK_RE.search(str(name)))
    if dtype == "object" or rate >= 1.0:
        conf = _CONFIRMED
    elif all_nan or (high_risk and n_failed > 0):
        conf = _HIGH_RISK
    else:
        conf = _AMBIGUOUS
    if all_nan:
        conf = _HIGH_RISK

    reason = (f"{dtype} at parse_rate={rate:.2f}" if dtype != "object"
              else "kept as text")
    if all_nan:
        reason = "all values failed coercion -> all-NaN"

    col_plan = {"dtype": dtype, "coerced_from": "object",
                "parse_rate": round(float(rate), 4), "n_failed": n_failed,
                "confidence": conf, "reason": reason, "suggest": suggest}
    return typed, col_plan


# ---------------------------------------------------------------------------
# Report rendering
# ---------------------------------------------------------------------------

def _report_frame(plan: dict, df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for col, cp in plan["columns"].items():
        s = df[col] if col in df.columns else pd.Series(dtype=object)
        flag = "" if cp["confidence"] == _CONFIRMED else "! "
        rows.append({
            "column": f"{flag}{col}",
            "dtype": cp["dtype"],
            "parsed_%": round(cp["parse_rate"] * 100, 2),
            "null_%": round(float(s.isna().mean()) * 100, 2) if len(s) else 0.0,
            "n_distinct": int(s.nunique(dropna=True)) if len(s) else 0,
            "problem": cp["reason"] if cp["confidence"] != _CONFIRMED else "",
            "action": cp["suggest"] or "",
            "confidence": cp["confidence"],
        })
    return pd.DataFrame(rows)


def _banner(plan: dict) -> str:
    d = plan["decisions"]
    src = plan["source"]
    if src.get("kind") == "sql":
        ro = d.get("read_only", {}).get("value")
        return (f"source={src['name']} | kind=sql | "
                f"read_only={bool(ro)} | "
                f"{plan['metadata']['n_rows']:,}x{plan['metadata']['n_cols']}")
    if src.get("kind") in ("parquet", "json"):
        extra = ""
        if src["kind"] == "json":
            jf = d.get("json_form", {})
            jfs = "" if jf.get("confidence") == _CONFIRMED else " (ambiguous)"
            extra = f" | form={jf.get('value')}{jfs}"
        return (f"source={src['name']} | kind={src['kind']}{extra} | "
                f"{plan['metadata']['n_rows']:,}x{plan['metadata']['n_cols']}")
    if src.get("kind") == "excel":
        sh, hdr = d["sheet"], d["header"]
        shs = "" if sh["confidence"] == _CONFIRMED else " (ambiguous)"
        depth = d.get("header_rows", {}).get("value", 1)
        plus = f" (+{depth - 1})" if depth > 1 else ""
        return (f"source={src['name']} | sheet={sh['value']!r}{shs} | "
                f"header=row {hdr['value']}{plus} | "
                f"{plan['metadata']['n_rows']:,}x{plan['metadata']['n_cols']}")
    enc, sep, hdr = d["encoding"], d["delimiter"], d["header"]
    encs = "" if enc["confidence"] == _CONFIRMED else " (ambiguous)"
    return (f"source={src['name']} | encoding={enc['value']}{encs} | "
            f"sep={sep['value']!r} | header=row {hdr['value']} | "
            f"{plan['metadata']['n_rows']:,}x{plan['metadata']['n_cols']}")


def _ambiguous_items(plan: dict):
    items = []
    for key, dec in plan["decisions"].items():
        if dec["confidence"] != _CONFIRMED:
            items.append(f"{key}={dec['value']!r} ({dec['reason']})")
    for col, cp in plan["columns"].items():
        if cp["confidence"] != _CONFIRMED:
            items.append(f"column {col!r}: {cp['confidence']} - {cp['reason']}")
    return items


def _decision_sentence(plan: dict) -> str:
    m = plan["metadata"]
    d = plan["decisions"]
    coerced = [(c, cp["dtype"]) for c, cp in plan["columns"].items()
               if cp["dtype"] != "object"]
    types = ", ".join(sorted({t for _, t in coerced})) or "none"
    n_failed = sum(cp["n_failed"] for cp in plan["columns"].values())
    a = m["n_ambiguous"]
    hint = " - re-run with params= to confirm" if a > 0 else ""
    if plan["source"].get("kind") == "sql":
        ro = d.get("read_only", {}).get("value")
        bracket = f"[sql, parametrised, read_only={bool(ro)}]"
    elif plan["source"].get("kind") == "parquet":
        bracket = "[parquet, typed]"
    elif plan["source"].get("kind") == "json":
        bracket = f"[json, form={d.get('json_form', {}).get('value')}]"
    elif plan["source"].get("kind") == "excel":
        depth = d.get("header_rows", {}).get("value", 1)
        bracket = (f"[sheet={d['sheet']['value']!r}, "
                   f"header=row {d['header']['value']}"
                   + (f" ({depth} rows)" if depth > 1 else "") + "]")
    else:
        bracket = (f"[encoding={d['encoding']['value']}, "
                   f"sep={d['delimiter']['value']!r}, "
                   f"header=row {d['header']['value']}]")
    return (f"Loaded {m['n_rows']:,} rows x {m['n_cols']} cols from "
            f"'{plan['source']['name']}' {bracket}; "
            f"coerced {len(coerced)} column(s) ({types}); {n_failed} cell(s) "
            f"failed -> NaN; {a} ambiguous decision(s){hint}. "
            f"Next: dx.audit(df).")


# ---------------------------------------------------------------------------
# Core build
# ---------------------------------------------------------------------------

def _parse_rows(text: str, sep: str):
    reader = csv.reader(io.StringIO(text), delimiter=sep, quotechar='"')
    return [r for r in reader]


def _parse_rows_iter(text: str, sep: str):
    """Lazy ``_parse_rows``: yield rows without building the full list."""
    return csv.reader(io.StringIO(text), delimiter=sep, quotechar='"')


def _build_from_text(text, source_meta, kind, on_ambiguous,
                     encoding_dec, sep_forced, header_forced, parse_dates,
                     decimal, thousands, na_values, max_rows, low_memory=False):
    sample_text = "\n".join(text.splitlines()[:200])
    dec0 = "." if decimal is None else decimal
    exclude = tuple(c for c in (dec0, thousands) if c and len(c) == 1)
    sep, sconf, sreason = _detect_delimiter(sample_text, sep_forced, exclude)
    if low_memory:
        # Stream the source once instead of materialising every cell as a
        # Python list-of-lists: keep only a bounded header sample plus a
        # per-row field-count tally (ints). Header detection and the ragged-
        # row disclosure below are computed from these, so the resulting frame
        # and plan are identical to the default path -- only the transient
        # peak drops (the larger of the two full-frame copies is never built).
        sample_rows, _all_lens = [], []
        for _r in _parse_rows_iter(text, sep):
            if not _r:
                continue
            _all_lens.append(len(_r))
            if len(sample_rows) < 50:
                sample_rows.append(_r)
        header_row, hconf, hreason = _detect_header(sample_rows, header_forced)
        data_counts = _all_lens[header_row + 1:]
    else:
        rows = [r for r in _parse_rows(text, sep) if r]
        header_row, hconf, hreason = _detect_header(rows, header_forced)
        data_counts = [len(r) for r in rows[header_row + 1:] if r]

    dec = "." if decimal is None else decimal
    extra_na = list(na_values) if na_values else []
    df = pd.read_csv(io.StringIO(text), sep=sep, header=header_row,
                     dtype=object, keep_default_na=True, na_values=extra_na,
                     skip_blank_lines=True, nrows=max_rows,
                     engine="python", on_bad_lines="skip")
    df.columns = [str(c) for c in df.columns]

    # ragged-row count computed above (default + low_memory paths)
    problems = []
    if data_counts:
        modal = max(set(data_counts), key=data_counts.count)
        ragged = sum(1 for c in data_counts if c != modal)
        if ragged:
            problems.append({"scope": "rows", "kind": "ragged",
                             "detail": f"{ragged} row(s) had != {modal} fields",
                             "action": "skipped on read"})

    columns, typed = {}, {}
    for col in df.columns:
        ts, cp = _infer_column(col, df[col], parse_dates, dec, thousands)
        typed[col] = ts
        columns[col] = cp
        if cp["confidence"] == _HIGH_RISK and "all-NaN" in cp["reason"]:
            problems.append({"scope": col, "kind": "all_nan",
                             "detail": cp["reason"], "action": "kept as NaN"})

    out = pd.DataFrame(typed)

    n_amb_cols = sum(1 for cp in columns.values() if cp["confidence"] != _CONFIRMED)
    n_amb_parse = sum(1 for c in (sconf, hconf, encoding_dec[1]) if c != _CONFIRMED)
    plan = {
        "function": "load",
        "source": {**source_meta, "kind": kind},
        "parse": {"encoding": encoding_dec[0], "delimiter": sep, "quotechar": '"',
                  "header_row": header_row, "skiprows": list(range(header_row)),
                  "decimal": dec, "thousands": thousands,
                  "na_values": extra_na},
        "columns": columns,
        "problems": problems,
        "decisions": {
            "encoding": {"value": encoding_dec[0], "confidence": encoding_dec[1],
                         "reason": encoding_dec[2]},
            "delimiter": {"value": sep, "confidence": sconf, "reason": sreason},
            "header": {"value": header_row, "confidence": hconf, "reason": hreason},
            "decimal": {"value": dec, "confidence": _CONFIRMED, "reason": "resolved"},
        },
        "policy": {"on_ambiguous": on_ambiguous,
                   "max_rows": max_rows},
        "metadata": {"n_rows": int(out.shape[0]), "n_cols": int(out.shape[1]),
                     "n_ambiguous": int(n_amb_cols + n_amb_parse)},
        "version": __version__,
        "generated_at": now_iso(),
    }
    return out, json_safe(plan)


def _apply_plan(text, plan, parse_dates):
    """Replay: apply a stored plan verbatim (no detection)."""
    p = plan["parse"]
    df = pd.read_csv(io.StringIO(text), sep=p["delimiter"], header=p["header_row"],
                     dtype=object, keep_default_na=True,
                     na_values=list(p.get("na_values") or []),
                     skip_blank_lines=True, nrows=plan["policy"].get("max_rows"),
                     engine="python", on_bad_lines="skip")
    df.columns = [str(c) for c in df.columns]
    typed = {}
    for col in df.columns:
        cp = plan["columns"].get(col, {"dtype": "object"})
        dtype = cp.get("dtype", "object")
        s = df[col]
        if dtype.startswith("datetime"):
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                typed[col] = pd.to_datetime(s, errors="coerce")
        elif dtype in ("float64", "int64"):
            ts, _, _ = _try_numeric(s, p.get("decimal", "."), p.get("thousands"))
            typed[col] = ts if ts is not None else pd.to_numeric(s, errors="coerce")
        elif dtype == "boolean":
            ts, _ = _try_bool(s)
            typed[col] = ts if ts is not None else s
        else:
            typed[col] = s
    return pd.DataFrame(typed)



# ---------------------------------------------------------------------------
# Excel (Phase 11.2): sheets / data block / multi-row headers / values only
# ---------------------------------------------------------------------------

def _require_openpyxl():
    try:
        import openpyxl  # lazy `io` extra
        return openpyxl
    except ImportError as exc:
        raise DextraLoaderError(
            "load: Excel sources need openpyxl. Install it with "
            '`pip install "dextra[io]"` (or `pip install openpyxl`).') from exc


def _cell_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _excel_sheet_rows(raw: bytes, sheet):
    """Open the workbook (cached values only -- never formulas, never macros)
    and pick a sheet. Returns ``(rows, name, sheets_meta, confidence, reason)``.
    """
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True,
                                    data_only=True)
    except Exception as exc:  # noqa: BLE001 - normalise to a loader error
        raise DextraLoaderError(
            f"load: could not open the Excel workbook ({exc}).") from exc
    try:
        names = wb.sheetnames
        sheets_meta = [{"name": ws.title,
                        "visible": ws.sheet_state == "visible"}
                       for ws in wb.worksheets]
        if isinstance(sheet, bool):
            raise TypeError("load: sheet must be a name (str) or index (int).")
        if sheet is None:
            visible = [m["name"] for m in sheets_meta if m["visible"]]
            candidates = visible or names
            name = candidates[0]
            if len(candidates) == 1:
                conf, reason = _CONFIRMED, f"single sheet ({name!r})"
            else:
                conf = _AMBIGUOUS
                reason = (f"{len(candidates)} sheets {candidates!r} -> "
                          f"defaulted to first visible {name!r}; pass sheet= "
                          "to choose")
        elif isinstance(sheet, int):
            if not 0 <= sheet < len(names):
                raise DextraLoaderError(
                    f"load: sheet index {sheet} out of range; workbook has "
                    f"{len(names)} sheet(s): {names!r}.")
            name = names[sheet]
            conf, reason = _CONFIRMED, f"user-specified (index {sheet} -> {name!r})"
        else:
            name = str(sheet)
            if name not in names:
                raise DextraLoaderError(
                    f"load: sheet {name!r} not found; available sheets: "
                    f"{names!r}.")
            conf, reason = _CONFIRMED, f"user-specified ({name!r})"
        rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
    finally:
        wb.close()
    return rows, name, sheets_meta, conf, reason


def _detect_block(rows):
    """Trim fully-empty border rows/columns; return (block, bounds, problems)."""
    row_used = [not all(_cell_empty(c) for c in r) for r in rows]
    if not any(row_used):
        return [], {"first_row": 0, "last_row": -1,
                    "first_col": 0, "last_col": -1}, []
    first_row = row_used.index(True)
    last_row = len(rows) - 1 - row_used[::-1].index(True)
    sub = rows[first_row:last_row + 1]
    width = max(len(r) for r in sub)
    sub = [list(r) + [None] * (width - len(r)) for r in sub]
    col_used = [any(not _cell_empty(r[j]) for r in sub) for j in range(width)]
    first_col = col_used.index(True)
    last_col = width - 1 - col_used[::-1].index(True)
    block = [r[first_col:last_col + 1] for r in sub]
    problems = []
    if first_row or first_col:
        problems.append({
            "scope": "sheet", "kind": "offset_block",
            "detail": (f"data block starts at row {first_row}, "
                       f"column {first_col} (0-based)"),
            "action": "leading empty rows/columns skipped"})
    return block, {"first_row": first_row, "last_row": last_row,
                   "first_col": first_col, "last_col": last_col}, problems


def _header_probe_rows(block):
    """Stringify block rows with trailing empties trimmed (for _detect_header)."""
    probe = []
    for r in block:
        idx = [j for j, c in enumerate(r) if not _cell_empty(c)]
        end = idx[-1] + 1 if idx else 0
        probe.append(["" if _cell_empty(c) else str(c) for c in r[:end]])
    return probe


def _detect_header_span(block, hdr, hconf, hreason, forced_rows):
    """Extend a detected header row into a multi-row header span.

    Merged-cell signature: a level above the current header carries labels
    that sit over *gaps* of the row below (the anchored cell of a merged
    span); a level below is a full text row under a gappy header. Returns
    ``(hdr, depth, confidence, reason)``; ``forced_rows`` forces the number
    of header rows counted downward from ``hdr``.
    """
    if forced_rows is not None:
        n = max(1, int(forced_rows))
        return hdr, n, _CONFIRMED, f"user-specified ({n} header row(s))"
    width = len(block[0]) if block else 0
    depth = 1
    # Upward: a partial label row whose cells cover gaps of the row below.
    while hdr > 0 and depth < 3:
        above = block[hdr - 1]
        cur = block[hdr]
        above_idx = [j for j, c in enumerate(above) if not _cell_empty(c)]
        cur_gaps = {j for j, c in enumerate(cur) if _cell_empty(c)}
        if (above_idx and len(above_idx) < width
                and any(j in cur_gaps for j in above_idx)):
            hdr -= 1
            depth += 1
        else:
            break
    # Downward: gaps (merged cells) in the bottom level + a full text row below.
    while depth < 3 and hdr + depth < len(block):
        bottom = block[hdr + depth - 1]
        nxt = block[hdr + depth]
        nxt_cells = [c for c in nxt if not _cell_empty(c)]
        nxt_texty = (bool(nxt_cells)
                     and all(not _looks_numeric(str(c))
                             and not isinstance(c, _dt.date)
                             for c in nxt_cells)
                     and len(nxt_cells) == width)
        if any(_cell_empty(c) for c in bottom) and nxt_texty:
            depth += 1
        else:
            break
    if depth == 1:
        return hdr, 1, hconf, hreason
    return hdr, depth, _AMBIGUOUS, (
        f"rows {hdr}..{hdr + depth - 1} look like one merged header -> "
        f"combined {depth} rows; pass header_rows= to override")


def _combine_headers(block, hdr, depth):
    """Build column names from ``depth`` header rows (upper levels ffilled)."""
    width = len(block[0])
    levels = []
    for k in range(depth):
        row = block[hdr + k] if hdr + k < len(block) else [None] * width
        if k < depth - 1:  # upper levels: forward-fill across merged spans
            filled, last = [], None
            for c in row:
                if not _cell_empty(c):
                    last = str(c).strip()
                filled.append(last)
            levels.append(filled)
        else:              # bottom level: taken as written
            levels.append([None if _cell_empty(c) else str(c).strip()
                           for c in row])
    names, seen = [], {}
    for j in range(width):
        parts = []
        for lev in levels:
            v = lev[j]
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        name = "_".join(parts) if parts else f"col{j}"
        k = seen.get(name, 0)
        seen[name] = k + 1
        if k:
            name = f"{name}.{k}"
        names.append(name)
    return names


def _type_excel_column(name, values, parse_dates, decimal, thousands,
                       na_tokens):
    """Type one Excel column: native cell types first, else measured inference."""
    vals = []
    for v in values:
        if isinstance(v, str):
            t = v.strip()
            vals.append(None if t == "" or t in na_tokens else v)
        else:
            vals.append(v)
    nn = [v for v in vals if v is not None]
    s = pd.Series(vals, dtype=object)

    def _native(dtype_series, coerced_from, reason):
        return dtype_series, {
            "dtype": str(dtype_series.dtype), "coerced_from": coerced_from,
            "parse_rate": 1.0, "n_failed": 0, "confidence": _CONFIRMED,
            "reason": reason, "suggest": None}

    if nn and all(isinstance(v, _dt.date) and not isinstance(v, bool)
                  for v in nn):
        if parse_dates:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                return _native(pd.to_datetime(s, errors="coerce"),
                               "excel-date", "native Excel date cells")
    elif nn and all(isinstance(v, bool) for v in nn):
        return _native(s.astype("boolean"), "excel-bool",
                       "native Excel boolean cells")
    elif nn and all(isinstance(v, (int, float)) and not isinstance(v, bool)
                    for v in nn):
        return _native(pd.to_numeric(s, errors="coerce"), "excel-number",
                       "native Excel numeric cells")
    as_str = pd.Series([None if v is None else str(v) for v in vals],
                       dtype=object)
    return _infer_column(name, as_str, parse_dates, decimal, thousands)


def _build_from_excel(raw, source_meta, on_ambiguous, sheet, header_forced,
                      header_rows_forced, parse_dates, decimal, thousands,
                      na_values, max_rows):
    rows, sheet_name, sheets_meta, sconf, sreason = _excel_sheet_rows(raw, sheet)
    block, bounds, problems = _detect_block(rows)
    dec = "." if decimal is None else decimal
    na_tokens = set(na_values or [])

    columns, typed = {}, {}
    if not block:
        hdr, depth = 0, 1
        hconf, hreason = _HIGH_RISK, "sheet is empty"
        problems.append({"scope": "sheet", "kind": "empty",
                         "detail": f"sheet {sheet_name!r} has no data",
                         "action": "returned an empty frame"})
    else:
        probe = _header_probe_rows(block)
        hdr0, hconf, hreason = _detect_header(probe, header_forced)
        hdr, depth, hconf, hreason = _detect_header_span(
            block, hdr0, hconf, hreason, header_rows_forced)
        names = _combine_headers(block, hdr, depth)
        data = block[hdr + depth:]
        if max_rows is not None:
            data = data[:max_rows]
        for j, name in enumerate(names):
            colvals = [r[j] for r in data]
            ts, cp = _type_excel_column(name, colvals, parse_dates, dec,
                                        thousands, na_tokens)
            typed[name] = ts
            columns[name] = cp
            if cp["confidence"] == _HIGH_RISK and "all-NaN" in cp["reason"]:
                problems.append({"scope": name, "kind": "all_nan",
                                 "detail": cp["reason"],
                                 "action": "kept as NaN"})
    out = pd.DataFrame(typed)

    n_amb_cols = sum(1 for cp in columns.values()
                     if cp["confidence"] != _CONFIRMED)
    n_amb_parse = sum(1 for c in (sconf, hconf) if c != _CONFIRMED)
    plan = {
        "function": "load",
        "source": {**source_meta, "kind": "excel"},
        "parse": {"sheet": sheet_name, "header_row": hdr,
                  "header_rows": depth, "block": bounds, "decimal": dec,
                  "thousands": thousands, "na_values": sorted(na_tokens)},
        "sheets": sheets_meta,
        "columns": columns,
        "problems": problems,
        "decisions": {
            "sheet": {"value": sheet_name, "confidence": sconf,
                      "reason": sreason},
            "header": {"value": hdr, "confidence": hconf, "reason": hreason},
            "header_rows": {"value": depth, "confidence": _CONFIRMED,
                            "reason": "part of the header decision"},
            "decimal": {"value": dec, "confidence": _CONFIRMED,
                        "reason": "resolved"},
        },
        "policy": {"on_ambiguous": on_ambiguous,
                   "max_rows": max_rows},
        "metadata": {"n_rows": int(out.shape[0]), "n_cols": int(out.shape[1]),
                     "n_ambiguous": int(n_amb_cols + n_amb_parse)},
        "version": __version__,
        "generated_at": now_iso(),
    }
    return out, json_safe(plan)


def _apply_plan_excel(raw, plan, parse_dates):
    """Replay: apply a stored Excel plan verbatim (no detection)."""
    p = plan["parse"]
    rows, _, _, _, _ = _excel_sheet_rows(raw, p.get("sheet"))
    block, _, _ = _detect_block(rows)
    hdr = int(p.get("header_row", 0))
    depth = int(p.get("header_rows", 1))
    names = list(plan["columns"].keys())
    data = block[hdr + depth:] if block else []
    max_rows = (plan.get("policy") or {}).get("max_rows")
    if max_rows is not None:
        data = data[:max_rows]
    na_tokens = set(p.get("na_values") or [])
    typed = {}
    for j, name in enumerate(names):
        colvals = [(r[j] if j < len(r) else None) for r in data]
        vals = [None if (isinstance(v, str)
                         and (v.strip() == "" or v.strip() in na_tokens))
                else v for v in colvals]
        s = pd.Series(vals, dtype=object)
        dtype = plan["columns"].get(name, {}).get("dtype", "object")
        if dtype.startswith("datetime"):
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                typed[name] = pd.to_datetime(s, errors="coerce")
        elif dtype in ("float64", "int64"):
            s_str = s.map(lambda v: v if v is None else str(v))
            ts, _, _ = _try_numeric(s_str, p.get("decimal", "."),
                                    p.get("thousands"))
            typed[name] = (ts if ts is not None
                           else pd.to_numeric(s_str, errors="coerce"))
        elif dtype == "boolean":
            ts, _ = _try_bool(s)
            typed[name] = ts if ts is not None else s
        else:
            # explicit object dtype: pandas 3.0 would otherwise infer str
            typed[name] = pd.Series(
                [None if v is None else str(v) for v in vals], dtype=object)
    return pd.DataFrame(typed)



# ---------------------------------------------------------------------------
# Parquet + JSON/NDJSON (Phase 11.3a)
# ---------------------------------------------------------------------------

def _have_parquet_engine() -> bool:
    """True when a parquet engine (pyarrow / fastparquet) is importable."""
    for mod in ("pyarrow", "fastparquet"):
        try:
            __import__(mod)
            return True
        except ImportError:
            continue
    return False


def _passthrough_columns(df, parse_dates, decimal, thousands):
    """Type columns of an already-typed frame: re-infer text, confirm the rest.

    Returns ``(typed_dict, columns_plan_dict)`` (shared by the in-memory
    pass-through, parquet, and replayed typed sources).
    """
    columns, typed = {}, {}
    for col in df.columns:
        dt_ = df[col].dtype
        if pd.api.types.is_object_dtype(dt_) or pd.api.types.is_string_dtype(dt_):
            ts, cp = _infer_column(col, df[col].astype(object), parse_dates,
                                   decimal, thousands)
        else:
            ts = df[col]
            cp = {"dtype": str(dt_), "coerced_from": str(dt_),
                  "parse_rate": 1.0, "n_failed": 0, "confidence": _CONFIRMED,
                  "reason": "already typed", "suggest": None}
        typed[str(col)] = ts
        columns[str(col)] = cp
    return typed, columns


def _read_parquet_frame(raw: bytes) -> pd.DataFrame:
    if not _have_parquet_engine():
        raise DextraLoaderError(
            "load: parquet needs an engine. Install one with "
            '`pip install "dextra[perf]"` (pyarrow).')
    try:
        return pd.read_parquet(io.BytesIO(raw))
    except DextraLoaderError:
        raise
    except Exception as exc:  # e.g. pyarrow.lib.ArrowInvalid (pyarrow is optional)
        raise DextraLoaderError(
            "load: could not read the data as parquet -- the bytes are not a "
            "valid parquet file. A wrong extension on a CSV / Excel / JSON "
            "file is the usual cause; pass an explicit kind= (e.g. "
            "kind='csv') if the content is not parquet. Underlying cause: "
            f"{type(exc).__name__}: {exc}") from exc


def _build_from_parquet(raw, source_meta, on_ambiguous, parse_dates,
                        decimal, thousands, max_rows):
    df = _read_parquet_frame(raw)
    if max_rows is not None:
        df = df.head(max_rows)
    dec = "." if decimal is None else decimal
    typed, columns = _passthrough_columns(df, parse_dates, dec, thousands)
    out = pd.DataFrame(typed)
    n_amb = sum(1 for cp in columns.values() if cp["confidence"] != _CONFIRMED)
    plan = {
        "function": "load",
        "source": {**source_meta, "kind": "parquet"},
        "parse": {"decimal": dec, "thousands": thousands},
        "columns": columns,
        "problems": [],
        "decisions": {"format": {"value": "parquet", "confidence": _CONFIRMED,
                                 "reason": "typed columnar source"}},
        "policy": {"on_ambiguous": on_ambiguous,
                   "max_rows": max_rows},
        "metadata": {"n_rows": int(out.shape[0]), "n_cols": int(out.shape[1]),
                     "n_ambiguous": int(n_amb)},
        "version": __version__,
        "generated_at": now_iso(),
    }
    return out, json_safe(plan)


def _json_records(text: str):
    """Parse JSON payload into records. Returns (records, form, conf, reason,
    n_bad) where form is 'records-array' / 'ndjson' / 'single-object'."""
    stripped = text.lstrip()
    if not stripped:
        return [], "empty", _HIGH_RISK, "empty JSON payload", 0
    if stripped[0] == "[":
        data = json.loads(text)
        if not isinstance(data, list):
            raise DextraLoaderError("load: JSON array expected at top level.")
        bad = sum(1 for r in data if not isinstance(r, dict))
        recs = [r for r in data if isinstance(r, dict)]
        if not recs:
            raise DextraLoaderError(
                "load: JSON array contains no object records.")
        return recs, "records-array", _CONFIRMED, "top-level JSON array", bad
    if stripped[0] == "{":
        lines = [ln for ln in text.splitlines() if ln.strip()]
        if len(lines) > 1:
            recs, bad = [], 0
            for ln in lines:
                try:
                    obj = json.loads(ln)
                    if isinstance(obj, dict):
                        recs.append(obj)
                    else:
                        bad += 1
                except json.JSONDecodeError:
                    bad += 1
            if recs and bad <= len(lines) // 2:
                return recs, "ndjson", _CONFIRMED, "one JSON object per line", bad
        obj = json.loads(text)  # single object -> one record (disclosed)
        return [obj], "single-object", _AMBIGUOUS, (
            "single top-level object -> loaded as one record"), 0
    raise DextraLoaderError(
        "load: unrecognised JSON payload (expected an array of objects, "
        "NDJSON lines, or a single object).")


def _build_from_json(text, source_meta, on_ambiguous, parse_dates,
                     decimal, thousands, na_values, max_rows):
    try:
        records, form, fconf, freason, n_bad = _json_records(text)
    except json.JSONDecodeError as exc:
        raise DextraLoaderError(f"load: invalid JSON ({exc}).") from exc
    if max_rows is not None:
        records = records[:max_rows]
    dec = "." if decimal is None else decimal
    na_tokens = set(na_values or [])

    names, nested = [], set()
    for r in records:
        for k in r:
            if str(k) not in names:
                names.append(str(k))
    cols = {n: [] for n in names}
    for r in records:
        for n in names:
            v = r.get(n)
            if isinstance(v, str) and (v.strip() == "" or v.strip() in na_tokens):
                v = None
            elif isinstance(v, (dict, list)):
                nested.add(n)
                v = json.dumps(v, ensure_ascii=False)
            cols[n].append(v)

    problems = []
    if n_bad:
        problems.append({"scope": "rows", "kind": "bad_records",
                         "detail": f"{n_bad} non-object / unparsable item(s)",
                         "action": "skipped"})
    for n in sorted(nested):
        problems.append({"scope": n, "kind": "nested",
                         "detail": "nested object/array values",
                         "action": "serialised to JSON strings"})

    columns, typed = {}, {}
    for n in names:
        ts, cp = _type_excel_column(n, cols[n], parse_dates, dec, thousands,
                                    na_tokens)
        typed[n] = ts
        columns[n] = cp
    out = pd.DataFrame(typed)

    n_amb_cols = sum(1 for cp in columns.values()
                     if cp["confidence"] != _CONFIRMED)
    n_amb = n_amb_cols + (0 if fconf == _CONFIRMED else 1)
    plan = {
        "function": "load",
        "source": {**source_meta, "kind": "json"},
        "parse": {"form": form, "decimal": dec, "thousands": thousands,
                  "na_values": sorted(na_tokens)},
        "columns": columns,
        "problems": problems,
        "decisions": {"json_form": {"value": form, "confidence": fconf,
                                    "reason": freason}},
        "policy": {"on_ambiguous": on_ambiguous,
                   "max_rows": max_rows},
        "metadata": {"n_rows": int(out.shape[0]), "n_cols": int(out.shape[1]),
                     "n_ambiguous": int(n_amb)},
        "version": __version__,
        "generated_at": now_iso(),
    }
    return out, json_safe(plan)


def _coerce_replay_column(s: pd.Series, dtype: str, parse_block: dict,
                          vals: list):
    """Coerce one object column to a stored plan dtype (replay; no detection)."""
    if dtype.startswith("datetime"):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return pd.to_datetime(s, errors="coerce")
    if dtype in ("float64", "int64"):
        s_str = pd.Series([None if v is None else str(v) for v in vals],
                          dtype=object)
        ts, _, _ = _try_numeric(s_str, parse_block.get("decimal", "."),
                                parse_block.get("thousands"))
        return ts if ts is not None else pd.to_numeric(s_str, errors="coerce")
    if dtype == "boolean":
        ts, _ = _try_bool(s)
        return ts if ts is not None else s
    if dtype == "object":
        return pd.Series([None if v is None else str(v) for v in vals],
                         dtype=object)
    return s  # already-typed dtypes pass through


def _apply_plan_parquet(raw, plan):
    df = _read_parquet_frame(raw)
    max_rows = (plan.get("policy") or {}).get("max_rows")
    if max_rows is not None:
        df = df.head(max_rows)
    p = plan.get("parse", {})
    typed = {}
    for col in df.columns:
        name = str(col)
        dtype = plan["columns"].get(name, {}).get("dtype", str(df[col].dtype))
        dt_ = df[col].dtype
        if pd.api.types.is_object_dtype(dt_) or pd.api.types.is_string_dtype(dt_):
            vals = [None if pd.isna(v) else v for v in df[col].tolist()]
            typed[name] = _coerce_replay_column(
                pd.Series(vals, dtype=object), dtype, p, vals)
        else:
            typed[name] = df[col]
    return pd.DataFrame(typed)


def _apply_plan_json(text, plan, parse_dates):
    p = plan.get("parse", {})
    records, _, _, _, _ = _json_records(text)
    max_rows = (plan.get("policy") or {}).get("max_rows")
    if max_rows is not None:
        records = records[:max_rows]
    na_tokens = set(p.get("na_values") or [])
    names = list(plan["columns"].keys())
    typed = {}
    for n in names:
        vals = []
        for r in records:
            v = r.get(n)
            if isinstance(v, str) and (v.strip() == "" or v.strip() in na_tokens):
                v = None
            elif isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            vals.append(v)
        dtype = plan["columns"].get(n, {}).get("dtype", "object")
        typed[n] = _coerce_replay_column(pd.Series(vals, dtype=object),
                                         dtype, p, vals)
    return pd.DataFrame(typed)



# ---------------------------------------------------------------------------
# Safe SQL (Phase 11.3b): parametrised only, read-only files, row guard
# ---------------------------------------------------------------------------

def _open_sql_source(source):
    """Return ``(conn, own_connection, display_name, read_only)``.

    A SQLite file path is opened **read-only** (URI ``mode=ro``); a caller
    DB-API connection is used as-is (read-only cannot be enforced and is
    disclosed as such in the plan).
    """
    if isinstance(source, (str, os.PathLike)):
        path = os.fspath(source)
        ext = os.path.splitext(path)[1].lower()
        if ext not in _SQLITE_EXT:
            raise DextraLoaderError(
                "load: sql= needs a SQLite file "
                "(.db / .sqlite / .sqlite3) or an open DB-API connection; "
                f"got {path!r}.")
        if not os.path.exists(path):
            raise DextraLoaderError(
                f"load: database file not found: {path!r}.")
        import sqlite3
        from urllib.parse import quote
        qp = quote(os.path.abspath(path).replace(os.sep, "/"), safe="/:")
        if not qp.startswith("/"):
            qp = "/" + qp
        conn = sqlite3.connect(f"file:{qp}?mode=ro", uri=True)
        return conn, True, os.path.basename(path), True
    if hasattr(source, "cursor"):
        return source, False, type(source).__name__, False
    raise DextraLoaderError(
        "load: sql= needs a SQLite file path or an open DB-API connection; "
        f"got {type(source).__name__}.")


def _execute_sql(conn, sql, sql_params, max_rows):
    """Run ONE parametrised statement; fetch at most the row guard + 1."""
    if not isinstance(sql, str) or not sql.strip():
        raise DextraLoaderError("load: sql must be a non-empty string.")
    if ";" in sql.strip().rstrip(";"):
        raise DextraLoaderError(
            "load: one SQL statement only (stacked statements are refused).")
    guard = _SQL_ROW_GUARD if max_rows is None else int(max_rows)
    cur = conn.cursor()
    try:
        try:
            cur.execute(sql, sql_params if sql_params is not None else ())
        except Exception as exc:
            raise DextraLoaderError(
                f"load: SQL execution failed ({exc}).") from exc
        if cur.description is None:
            raise DextraLoaderError(
                "load: the SQL statement returned no result set "
                "(a SELECT is expected).")
        names = [str(d[0]) for d in cur.description]
        rows = cur.fetchmany(guard + 1)
        truncated = len(rows) > guard
        if truncated:
            rows = rows[:guard]
    finally:
        cur.close()
    return names, rows, truncated, guard


def _build_from_sql(names, rows, truncated, guard, sql, sql_params,
                    source_name, read_only, on_ambiguous, parse_dates,
                    decimal, thousands, na_values, max_rows):
    dec = "." if decimal is None else decimal
    na_tokens = set(na_values or [])
    n_params = len(sql_params) if sql_params else 0

    columns, typed = {}, {}
    for j, name in enumerate(names):
        colvals = [r[j] for r in rows]
        ts, cp = _type_excel_column(name, colvals, parse_dates, dec,
                                    thousands, na_tokens)
        typed[name] = ts
        columns[name] = cp
    out = pd.DataFrame(typed) if typed else pd.DataFrame()

    problems = []
    guard_conf, guard_reason = _CONFIRMED, "not reached"
    if truncated:
        user_cap = max_rows is not None
        problems.append({
            "scope": "rows", "kind": "row_guard",
            "detail": f"result truncated at {guard:,} row(s)",
            "action": ("user max_rows= cap" if user_cap else
                       "default row guard; pass max_rows= to raise it")})
        if not user_cap:
            guard_conf = _AMBIGUOUS
            guard_reason = f"default guard hit at {guard:,} rows"
        else:
            guard_reason = f"user cap hit at {guard:,} rows"

    n_amb_cols = sum(1 for cp in columns.values()
                     if cp["confidence"] != _CONFIRMED)
    n_amb = n_amb_cols + (0 if guard_conf == _CONFIRMED else 1)
    plan = {
        "function": "load",
        "source": {"name": source_name, "kind": "sql", "sha256": None,
                   "size": None, "mtime": None},
        "parse": {"sql": sql, "sql_params": sql_params,
                  "read_only": bool(read_only), "decimal": dec,
                  "thousands": thousands, "na_values": sorted(na_tokens)},
        "columns": columns,
        "problems": problems,
        "decisions": {
            "sql": {"value": "parametrised", "confidence": _CONFIRMED,
                    "reason": (f"{n_params} bound parameter(s)" if n_params
                               else "no parameters")},
            "read_only": {"value": bool(read_only), "confidence": _CONFIRMED,
                          "reason": ("sqlite opened with mode=ro" if read_only
                                     else "caller-managed connection "
                                          "(not enforced)")},
            "row_guard": {"value": guard, "confidence": guard_conf,
                          "reason": guard_reason},
        },
        "policy": {"on_ambiguous": on_ambiguous,
                   "max_rows": max_rows},
        "metadata": {"n_rows": int(out.shape[0]), "n_cols": int(out.shape[1]),
                     "n_ambiguous": int(n_amb)},
        "version": __version__,
        "generated_at": now_iso(),
    }
    return out, json_safe(plan)


def _load_sql(source, *, sql, sql_params, params, on_ambiguous, parse_dates,
              decimal, thousands, na_values, max_rows, return_params, show,
              decimals, interactive):
    replaying = params is not None
    if replaying:
        p = params.get("parse", {})
        sql = p.get("sql")
        sql_params = p.get("sql_params")
        if max_rows is None:
            max_rows = (params.get("policy") or {}).get("max_rows")
    if sql is None:
        raise DextraLoaderError(
            "load: a database source needs sql= (parametrised; pass values "
            "via sql_params=, never by string formatting).")
    conn, own, name, read_only = _open_sql_source(source)
    try:
        names, rows, truncated, guard = _execute_sql(
            conn, sql, sql_params, max_rows)
    finally:
        if own:
            conn.close()

    if replaying:
        dec_block = params.get("parse", {})
        na_tokens = set(dec_block.get("na_values") or [])
        typed = {}
        for j, cname in enumerate(list(params["columns"].keys())):
            colvals = [(r[j] if j < len(r) else None) for r in rows]
            vals = [None if (isinstance(v, str)
                             and (v.strip() == "" or v.strip() in na_tokens))
                    else v for v in colvals]
            dtype = params["columns"].get(cname, {}).get("dtype", "object")
            typed[cname] = _coerce_replay_column(
                pd.Series(vals, dtype=object), dtype, dec_block, vals)
        out = pd.DataFrame(typed) if typed else pd.DataFrame()
        out.attrs = {}
        append_audit(out, {"stage": "loader", "function": "load",
                           "timestamp": now_iso(), "params": params,
                           "decision": "Replayed a stored SQL load plan."})
        if show:
            print(f"Decision: Replayed a stored SQL load plan on {name!r}.")
        return (out, params) if return_params else out

    out, plan = _build_from_sql(
        names, rows, truncated, guard, sql, sql_params, name, read_only,
        on_ambiguous, parse_dates, decimal, thousands, na_values, max_rows)
    return _disclose_and_finish(out, plan, on_ambiguous, show, decimals,
                                interactive, return_params)


def _disclose_and_finish(out, plan, on_ambiguous, show, decimals,
                         interactive, return_params):
    """Shared disclosure tail: report -> policy -> audit -> return."""
    items = _ambiguous_items(plan)
    if show:
        print(_banner(plan))
        with pd.option_context("display.max_columns", None,
                               "display.width", 0,
                               "display.float_format",
                               lambda v: f"{v:,.{decimals}f}"):
            print(_report_frame(plan, out).to_string(index=False))

    if on_ambiguous == "plan":
        if show:
            print(f"Decision: {_decision_sentence(plan)}")
        return plan

    if items and on_ambiguous == "raise":
        raise LoaderAmbiguityError(
            "ambiguous load decision(s):\n  - " + "\n  - ".join(items)
            + "\nOverride explicitly (encoding=/sep=/header_row=/dtype) or use "
              "on_ambiguous='warn'.")
    if items and on_ambiguous == "warn":
        warnings.warn("load: ambiguous decision(s): " + "; ".join(items),
                      DextraLoaderWarning, stacklevel=3)

    if interactive and show:
        resp = input("Apply this plan? [y]/abort: ").strip().lower()
        if resp == "abort":
            raise LoaderAbort("user aborted the load.")

    out.attrs = {}
    append_audit(out, {"stage": "loader", "function": "load",
                       "timestamp": plan["generated_at"], "params": plan,
                       "decision": _decision_sentence(plan)})
    if show:
        print(f"Decision: {_decision_sentence(plan)}")
    return (out, plan) if return_params else out


# ---------------------------------------------------------------------------
# Public API
# ===========================================================================
# 11  load  --  smart, transparent, replayable data loader
# ===========================================================================

def load(
    source,
    *,
    kind: str = "auto",
    params: Optional[dict] = None,
    on_ambiguous: str = "warn",
    encoding: Optional[str] = None,
    sep: Optional[str] = None,
    header_row: Optional[int] = None,
    sheet=None,
    header_rows: Optional[int] = None,
    sql: Optional[str] = None,
    sql_params=None,
    parse_dates: bool = True,
    decimal: Optional[str] = None,
    thousands: Optional[str] = None,
    na_values=None,
    max_rows: Optional[int] = None,
    low_memory: bool = False,
    sample_bytes: int = 262144,
    return_params: bool = False,
    show: bool = True,
    decimals: int = 4,
    df_name: Optional[str] = None,
    interactive: bool = False,
):
    """Load a messy source into a typed DataFrame, disclosing every decision.

    Auto-detects encoding, delimiter and the real header row of a delimited-text
    source, then infers per-column types and **measures** how many cells parsed.
    Transparency scales with uncertainty: confident parses load in one line;
    ambiguous decisions are flagged and handled per ``on_ambiguous``
    (``'warn'`` -> load + warn; ``'raise'`` -> raise; ``'plan'`` -> return the
    plan without loading). Every load emits a JSON-safe, replayable load plan
    (the ``params`` artifact); ``load(source, params=plan)`` reproduces the frame
    exactly. The source is never modified.

    Parameters
    ----------
    source : str | os.PathLike | file-like | DataFrame
        A path, an open binary/text stream, or an in-memory frame (typed
        pass-through). Excel = ``.xlsx`` / ``.xlsm`` (legacy ``.xls`` is
        refused with guidance); parquet = ``.parquet`` / ``.pq`` (typed,
        needs a lazy engine such as pyarrow); JSON = ``.json`` / ``.jsonl`` /
        ``.ndjson`` (records array or one object per line; nested values are
        serialised + disclosed). ``.pkl`` is always refused (pickle can
        execute arbitrary code on load).
    kind : {"auto", "csv", "tsv", "excel", "parquet", "json"}
        Source kind; inferred from the extension when ``"auto"``.
    params : dict, optional
        A previously returned load plan to replay deterministically.
    on_ambiguous : {"warn", "raise", "plan"}
        Policy when a decision is ambiguous (see above).
    encoding, sep, header_row, decimal, thousands : optional
        Force a decision instead of detecting it. ``encoding=`` is the
        deterministic way out for legacy text: auto-detection of non-utf-8
        bytes is a guess and is always flagged ambiguous (e.g. pass
        ``encoding="cp1256"`` for legacy Arabic files).
    sheet : str | int, optional
        Excel only: sheet name or 0-based index. Default: the single sheet,
        or the first visible one (flagged ambiguous when several exist).
    header_rows : int, optional
        Excel only: force the number of header rows (multi-row headers are
        otherwise detected and combined into ``top_bottom`` names).
    sql : str, optional
        SQL only: ONE parametrised statement (stacked statements are
        refused). Bind values with ``sql_params=`` -- never by string
        formatting. The source must be a SQLite file (``.db`` / ``.sqlite`` /
        ``.sqlite3``, opened **read-only**) or an open DB-API connection
        (read-only not enforceable; disclosed in the plan). Results are
        capped by ``max_rows`` or a default 1,000,000-row guard (truncation
        is disclosed and flagged).
    sql_params : dict | tuple, optional
        Named (``:name`` + dict) or positional (``?`` + tuple) parameters.
    parse_dates : bool
        Attempt safe datetime inference.
    na_values : list, optional
        Extra NA tokens added to the pandas defaults.
    max_rows : int, optional
        Safety cap on the number of rows read.
    low_memory : bool
        Delimited text only: stream the source once for header / ragged-row
        detection instead of materialising every cell as a Python
        list-of-lists, cutting the transient peak memory of a large CSV load
        (about a third on a million rows). Returns an identical frame and plan --
        per-column type inference is unchanged. A no-op for already-typed
        sources (Excel / parquet / JSON / SQL).
    return_params : bool
        Also return the load plan.
    show : bool
        Print the disclosure report + the ``Decision:`` sentence.
    decimals : int
        Numeric precision in the printed report.
    df_name : str, optional
        Name used in the audit / decision (inferred when omitted).
    interactive : bool
        Prompt to confirm before loading (classroom use; never during
        composition and never when ``show=False``).

    Returns
    -------
    pandas.DataFrame, or (DataFrame, plan) when ``return_params=True``; the plan
    alone when ``on_ambiguous='plan'``.

    Examples
    --------
    >>> df = dx.load('messy.csv')
    >>> df, plan = dx.load('messy.csv', return_params=True)
    >>> df = dx.load('messy.csv', params=plan)         # deterministic replay
    >>> df = dx.load('book.xlsx', sheet='Q1')          # Excel: pick a sheet
    >>> df = dx.load('app.db', sql='SELECT * FROM t WHERE d = :d',
    ...              sql_params={'d': '2026-01-01'})   # safe SQL
    """
    if on_ambiguous not in ("warn", "raise", "plan"):
        raise ValueError("on_ambiguous must be 'warn', 'raise' or 'plan'.")

    # In-memory frame -> typed pass-through.
    if isinstance(source, pd.DataFrame) or (
            not isinstance(source, (str, os.PathLike))
            and not hasattr(source, "read") and hasattr(source, "to_pandas")):
        return _load_frame(source, on_ambiguous=on_ambiguous,
                           parse_dates=parse_dates, decimal=decimal,
                           thousands=thousands, return_params=return_params,
                           show=show, df_name=df_name)

    # SQL sources (Phase 11.3b) -- resolved before any byte reading.
    is_sqlite_path = (isinstance(source, (str, os.PathLike)) and
                      os.path.splitext(os.fspath(source))[1].lower()
                      in _SQLITE_EXT)
    is_db_conn = hasattr(source, "cursor") and not hasattr(source, "read")
    if sql is not None or is_sqlite_path or is_db_conn:
        return _load_sql(source, sql=sql, sql_params=sql_params,
                         params=params, on_ambiguous=on_ambiguous,
                         parse_dates=parse_dates, decimal=decimal,
                         thousands=thousands, na_values=na_values,
                         max_rows=max_rows, return_params=return_params,
                         show=show, decimals=decimals,
                         interactive=interactive)

    if df_name is None:
        df_name = get_variable_name(source, depth=2)
        if df_name == "DataFrame":
            df_name = None

    raw, source_meta = _read_bytes(source)
    real_name = source_meta["name"]  # kind detection uses the real file name
    if df_name:
        source_meta = {**source_meta, "name": df_name}

    ext = os.path.splitext(real_name)[1].lower()
    if ext in _XLS_EXT:
        raise DextraLoaderError(
            "load: legacy .xls is not supported; save the file as .xlsx "
            "(or read it with pandas.read_excel + xlrd) and retry.")
    if kind == "auto":
        kind = ("excel" if ext in _EXCEL_EXT
                else "parquet" if ext in _PARQUET_EXT
                else "json" if ext in _JSON_EXT
                else "tsv" if ext in _TSV_EXT else "csv")
    if kind not in ("csv", "tsv", "excel", "parquet", "json"):
        raise ValueError("load: kind must be 'auto', 'csv', 'tsv', 'excel', "
                         "'parquet' or 'json'.")
    if kind == "tsv" and sep is None:
        sep = "\t"

    if kind not in ("excel", "parquet"):
        enc, econf, ereason = _detect_encoding(
            raw[:sample_bytes], encoding, truncated=len(raw) > sample_bytes)
        text, econf, ereason = _decode_full(raw, enc, econf, ereason)

    if params is not None:
        if kind == "excel":
            out = _apply_plan_excel(raw, params, parse_dates)
        elif kind == "parquet":
            out = _apply_plan_parquet(raw, params)
        elif kind == "json":
            out = _apply_plan_json(text, params, parse_dates)
        else:
            out = _apply_plan(text, params, parse_dates)
        src_sha = source_meta.get("sha256")
        plan_sha = params.get("source", {}).get("sha256")
        if src_sha and plan_sha and src_sha != plan_sha:
            entry = {"scope": "source", "kind": "source_changed",
                     "detail": "sha256 differs from the plan", "action": "replayed anyway"}
            params.setdefault("problems", []).append(entry)
            if on_ambiguous == "raise":
                raise LoaderReplayError(
                    "source has changed since the plan was created (sha256 mismatch).")
            warnings.warn("load: source changed since the plan was created.",
                          DextraLoaderWarning, stacklevel=2)
        out = out.copy()
        out.attrs = {}
        append_audit(out, {"stage": "loader", "function": "load",
                           "timestamp": now_iso(), "params": params,
                           "decision": "Replayed a stored load plan."})
        if show:
            print("Decision: Replayed a stored load plan on "
                  f"'{source_meta['name']}'.")
        return (out, params) if return_params else out

    if kind == "excel":
        out, plan = _build_from_excel(
            raw, source_meta, on_ambiguous, sheet, header_row, header_rows,
            parse_dates, decimal, thousands, na_values, max_rows)
    elif kind == "parquet":
        out, plan = _build_from_parquet(
            raw, source_meta, on_ambiguous, parse_dates,
            decimal, thousands, max_rows)
    elif kind == "json":
        out, plan = _build_from_json(
            text, source_meta, on_ambiguous, parse_dates,
            decimal, thousands, na_values, max_rows)
    else:
        out, plan = _build_from_text(
            text, source_meta, kind, on_ambiguous,
            (enc, econf, ereason), sep, header_row, parse_dates,
            decimal, thousands, na_values, max_rows, low_memory)

    return _disclose_and_finish(out, plan, on_ambiguous, show, decimals,
                                interactive, return_params)


def _load_frame(source, *, on_ambiguous, parse_dates, decimal, thousands,
                return_params, show, df_name):
    """Typed pass-through for an in-memory frame: re-infer object columns."""
    df = _ensure_pandas(source)
    name = df_name or get_variable_name(source, depth=3)
    dec = "." if decimal is None else decimal
    columns, typed = {}, {}
    for col in df.columns:
        dt = df[col].dtype
        if pd.api.types.is_object_dtype(dt) or pd.api.types.is_string_dtype(dt):
            ts, cp = _infer_column(col, df[col], parse_dates, dec, thousands)
        else:
            ts = df[col]
            cp = {"dtype": str(df[col].dtype), "coerced_from": str(df[col].dtype),
                  "parse_rate": 1.0, "n_failed": 0, "confidence": _CONFIRMED,
                  "reason": "already typed", "suggest": None}
        typed[str(col)] = ts
        columns[str(col)] = cp
    out = pd.DataFrame(typed)
    out.attrs = dict(df.attrs)
    n_amb = sum(1 for cp in columns.values() if cp["confidence"] != _CONFIRMED)
    plan = json_safe({
        "function": "load", "source": {"name": name, "kind": "frame",
                                        "sha256": None, "size": None, "mtime": None},
        "parse": {}, "columns": columns, "problems": [],
        "decisions": {}, "policy": {"on_ambiguous": on_ambiguous},
        "metadata": {"n_rows": int(out.shape[0]), "n_cols": int(out.shape[1]),
                     "n_ambiguous": int(n_amb)},
        "version": __version__, "generated_at": now_iso(),
    })
    decision = (f"Typed pass-through of in-memory frame '{name}': "
                f"{out.shape[0]:,} rows x {out.shape[1]} cols; "
                f"{n_amb} ambiguous column(s). Next: dx.audit(df).")
    append_audit(out, {"stage": "loader", "function": "load",
                       "timestamp": plan["generated_at"], "params": plan,
                       "decision": decision})
    if show:
        print(f"Decision: {decision}")
    return (out, plan) if return_params else out


def peek(source, *, kind: str = "auto", on_ambiguous: str = "plan",
         show: bool = True, n_preview: int = 10,
         df_name: Optional[str] = None, **load_kwargs):
    """Propose a load plan + preview WITHOUT committing a full load.

    Returns the load plan (dict); loads at most ``n_preview`` rows for the sample.
    The teaching / inspection entry point ("look before you load").
    ``df_name=`` labels the source explicitly in the plan / audit (audit #11);
    when omitted it is inferred as a last resort.
    """
    load_kwargs.pop("return_params", None)
    load_kwargs.pop("max_rows", None)
    plan = load(source, kind=kind, on_ambiguous="plan", show=show,
                max_rows=n_preview, df_name=df_name, **load_kwargs)
    return plan


# Short aliases, consistent with the underscore-free Phase 8/9/10 naming.
dload = load
dpeek = peek
